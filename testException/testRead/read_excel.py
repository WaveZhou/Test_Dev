
from openpyxl import load_workbook
import xlwings as xw
import xlrd,os
# wb = xw.Book("SL0771_久铭稳健7号私募证券投资基金资产估值表_2021-11-23.xls")
# sht = wb.sheets["sheet1"]
# name = sht.fullname
# data = sht.range('B28').value()
# print(data)
file_name = 'SL0771_久铭稳健7号私募证券投资基金资产估值表_2021-11-23.xlsx'
file_path = os.path.join(os.getcwd(),file_name)
xl = xlrd.open_workbook(file_path)
table = xl.sheets()[0]
res = table.cell(27,1).value
print(res)
class Read_Excel:
    def __init__(self,data_file):
        self.data_file = data_file
    def get_xlrd(self,row ,col):
         wb = load_workbook(self.data_file, data_only=True)
         ws = wb.worksheets[0]
         return ws.cell(row ,col).value

if __name__ == '__main__':
    re = Read_Excel('./SL0771_久铭稳健7号私募证券投资基金资产估值表_2021-11-23.xlsx')
    res = re.get_xlrd(30,1)
    print(res)