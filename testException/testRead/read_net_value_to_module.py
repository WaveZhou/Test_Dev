import openpyxl,os

base_dir = r'D:\Users\Administrator\AppData\Local\Programs\PycharmWorkSpace\Test_Dev\testException\origin'
#file = os.path.join(base_dir,'产品净值导入模板-20220909120107.xlsx')
work_book_obj = openpyxl.load_workbook(filename='产品净值导入模板-20220909120107.xls')
sheet = work_book_obj.sheetnames[0]
print(sheet)
# with open('产品净值导入模板-20220909120107.xlsx','r',encoding= 'charmap') as csvfile:
#     print(csvfile)
wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = '未到对账单账户信息'
# sheet['A1'] = '产品名称'
# sheet['B1'] = '券商'
# sheet['C1'] = '账户类型'
# sheet['D1'] = '账户号'
# sheet['E1'] = '上海卡号'
# sheet['F1'] = '深圳卡号'
# sheet['G1'] = '开户日期'
# sheet['H1'] = '所属营业部'
# sheet['I1'] = '联系人'
# sheet['J1'] = '邮箱'
# sheet['K1'] = '手机'
# sheet['L1'] = '电话'
# sheet['M1'] = '微信'